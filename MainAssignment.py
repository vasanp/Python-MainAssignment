import sys
import openpyxl
import json
from openpyxl import Workbook
from json import JSONEncoder

# Data file location
TestDataLoc = "/Users/vasanp/PycharmProjects/pythonProject/Python-Assignments/DataSheet.xlsx"

# Create a workbook Object
data_obj = openpyxl.load_workbook(TestDataLoc)

# Get Sheet
sheet_obj = data_obj["Details"]

movieDetails = ["Movie Details", "title", "genre", "length", "cast", "director", "rating", "language", "timing",
                "shows per day", "first show", "interval time", "movie gap", "capacity"]

# Store new users in a list
Users = []

# List of static timings
timings = ["10:00-12:00", "12:30-02:30", "03:00-05:00"]


# home page screen
def home():
    print("******Welcome to BookMyShow*******")
    print("1. Register New User ")
    print("2. Login ")
    print("3. Exit ")
    entry_action = int(input("choose your action: "))
    if entry_action == 1:
        add_new_user()
    elif entry_action == 2:
        login()
    elif entry_action == 3:
        exit()
    else:
        print("Choose a valid option")
        home()


# template for new users
class register:
    def __init__(self, username, user_email, user_phone, user_age, user_password):
        self.username = username
        self.user_email = user_email
        self.user_phone = user_phone
        self.user_age = user_age
        self.user_password = user_password


# Registering New Users
def add_new_user():
    print("****Create new Account*****")
    user_name = input("Name:")
    user_email = input("Email:")
    user_phone = input("Phone:")
    user_age = input("Age:")
    user_password = input("Password:")

    # getting last row in excel sheet
    last_row_user = sheet_obj.max_row

    # storing username and password for new users in excel(users) sheet
    sheet_obj.cell(row=last_row_user + 1, column=1).value = user_name
    sheet_obj.cell(row=last_row_user + 1, column=2).value = user_password

    # saving the excel
    data_obj.save('DataSheet.xlsx')

    # Creating instance of register class and storing in Users list
    Users.append(register(user_name, user_email, user_phone, user_age, user_password))

    # go to homepage
    home()


# template for admin role
class admin:
    admin_username = "admin1"
    admin_password = "password1"


# creating object of admin class
admin_obj = admin()


# tempplate for movie
class movie:
    def __init__(self, title, genre, length, cast, director, rating, language, shows, firstshow, intervaltime, moviegap,
                 capacity):
        self.title = title
        self.genre = genre
        self.length = length
        self.cast = cast
        self.director = director
        self.rating = rating
        self.language = language
        self.shows = shows
        self.firstshow = firstshow
        self.intervaltime = intervaltime
        self.moviegap = moviegap
        self.capacity = capacity


# Creating sheet/Intial Structure for storing all movies in excel
def excel_format(movie_title):
    data_obj.create_sheet(movie_title)
    worksheet = data_obj[movie_title]
    worksheet.cell(row=1, column=1).value = "title"
    worksheet.cell(row=1, column=2).value = "genre"
    worksheet.cell(row=1, column=3).value = "length"
    worksheet.cell(row=1, column=4).value = "cast"
    worksheet.cell(row=1, column=5).value = "director"
    worksheet.cell(row=1, column=6).value = "rating"
    worksheet.cell(row=1, column=7).value = "language"
    worksheet.cell(row=1, column=8).value = "timings"
    worksheet.cell(row=1, column=9).value = "shows_per_day"
    worksheet.cell(row=1, column=10).value = "firstshow"
    worksheet.cell(row=1, column=11).value = "interval_time"
    worksheet.cell(row=1, column=12).value = "gap"
    worksheet.cell(row=1, column=13).value = "capacity"
    worksheet.cell(row=1, column=14).value = "user_count"
    worksheet.cell(row=2, column=14).value = 1
    data_obj.save('DataSheet.xlsx')


# Login functionality
def login():
    print("******Welcome to BookMyShow******* ")
    user = input("User:")
    password = input("Password:")

    # Validating if creds matches with admin, route to admin functionalities else user functionalities
    if (user == admin_obj.admin_username and password == admin_obj.admin_password):
        admin_portal()
    else:
        login_validation(user, password)


# User credentials validation
def login_validation(user, password):
    # Iterating through excel rows to match user's username and password
    for i in range(1, sheet_obj.max_row + 1):
        username_validation = sheet_obj.cell(row=i, column=1)
        if username_validation.value == user:
            break

    user_password_validation = sheet_obj.cell(row=i, column=2)
    if user_password_validation.value == password:
        user_login_page(username_validation.value)
    else:
        # if credentials incorrect, route to home page
        home()


# Admin functionalities
def admin_portal():
    # Showing options in the admin portal
    print("******Welcome Admin*******")
    print("1.Add New Movie Info ")
    print("2.Edit Movie Info ")
    print("3.Delete Movies")
    print("4.Logout")

    # input for admin to perform action
    selected_admin_functionality = int(input("choose your admin action: "))

    if (selected_admin_functionality == 1):
        add_new_movie()
    elif (selected_admin_functionality == 2):
        edit_movie()
    elif (selected_admin_functionality == 3):
        delete_movie()
    elif (selected_admin_functionality == 4):
        home()
    else:
        print("Choose a valid option")
        home()


# Adding New Movie
def add_new_movie():
    # input for movie details
    print("******Welcome Admin*******")
    movie_title = input("Title  :")
    movie_genre = input("Genre  :")
    movie_length = input("Length  :")
    movie_cast = input("Cast  :")
    movie_director = input("Director  :")
    movie_rating = input("Admin rating  :")
    movie_language = input("Language:")
    movie_shows_per_day = input("Number of Shows in a day  :")
    movie_first_show = input("First Show  :")
    movie_interval_time = input("Interval Time  :")
    movie_gap = input("Gap Between Shows  :")
    movie_capacity = input("Capacity  :")
    print("New Movie Added!")
    # inserting all movies details into excel, movie sheet
    excel_format(movie_title)
    insert_movie_details_into_excel(movie_title, movie_genre, movie_length, movie_cast, movie_director, movie_rating,
                                    movie_language, movie_shows_per_day, movie_first_show, movie_interval_time,
                                    movie_gap, movie_capacity)
    # routing to admin portal
    admin_portal()


# logic for inserting movies into excel
def insert_movie_details_into_excel(movie_title, movie_genre, movie_length, movie_cast, movie_director, movie_rating,
                                    movie_language, movie_shows_per_day, movie_first_show, movie_interval_time,
                                    movie_gap, movie_capacity):
    # selecting the sheet and getting the last filled row
    worksheet = data_obj[movie_title]
    shows_per_day = int(movie_shows_per_day) + 2
    # inserting data into the cells and saving the excel
    for i in range(2, shows_per_day):
        worksheet.cell(row=i, column=1).value = movie_title
        worksheet.cell(row=i, column=2).value = movie_genre
        worksheet.cell(row=i, column=3).value = movie_length
        worksheet.cell(row=i, column=4).value = movie_cast
        worksheet.cell(row=i, column=5).value = movie_director
        worksheet.cell(row=i, column=6).value = movie_rating
        worksheet.cell(row=i, column=7).value = movie_language
        worksheet.cell(row=i, column=9).value = movie_shows_per_day
        worksheet.cell(row=i, column=10).value = movie_first_show
        worksheet.cell(row=i, column=11).value = movie_interval_time
        worksheet.cell(row=i, column=12).value = movie_gap
        worksheet.cell(row=i, column=13).value = movie_capacity
        insert_timings(timings, movie_title)
        data_obj.save('DataSheet.xlsx')


# insert timings in excel movie sheet
def insert_timings(timings, movie_title):
    worksheet = data_obj[movie_title]
    timings_length = len(timings)
    last_row_used = worksheet.max_row + 1
    try:
        for j in range(0, worksheet.max_row):
            time_val1 = worksheet.cell(row=j + 2, column=8)
            time_val1.value = timings[j]
    except IndexError:
        print("There can only be 3 shows")

    data_obj.save('DataSheet.xlsx')


# show movie timings
def show_movie_timings(user_selected_movie):
    worksheet = data_obj[user_selected_movie]
    last_row8 = worksheet.max_row
    index = 1
    for i in range(0, last_row8 - 1):
        time_val2 = worksheet.cell(row=i + 2, column=8)
        print(index, ".", time_val2.value)
        index = index + 1


# show movie details
def show_movie_details(movie_details):
    worksheet = data_obj[movie_details]
    for x in range(1, len(movieDetails)):
        print(movieDetails[x], ": ", worksheet.cell(row=2, column=x).value)


# Edit movie page
def edit_movie():
    print("******Welcome Admin*******")
    print("Select movie which you want to edit:")
    # show movie names
    show_movie_name()
    # movie_name()
    movie_to_edit = input(">>>")

    # calling edit movie function
    edit_movie_details(movie_to_edit)


# Edit functionality abstraction
def edit_movie_details(movie_to_edit):
    show_movie_details(movie_to_edit)
    print("Please write the option what needs to be changed. For ex - title ")
    movie_field_update = input(">>>")
    print("Please enter new value for the above selected option")
    movie_field_new_value = input(">>>")
    edit_movie_excel(movie_field_update, movie_to_edit, movie_field_new_value)


# Code for fetching which field to update in excel
def edit_movie_excel(movie_field_update, movie_to_edit, movie_field_new_value):
    worksheet = data_obj[movie_to_edit]
    index = movieDetails.index(movie_field_update)
    last_row8 = worksheet.max_row + 1
    for i in range(2, last_row8):
        worksheet.cell(row=i, column=index).value = movie_field_new_value

    if movie_field_update == "title":
        worksheet.title = movie_field_new_value
    data_obj.save('DataSheet.xlsx')
    show_movie_details(worksheet.title)
    admin_portal()


# delete movie page
def delete_movie():
    print("******Welcome Admin*******")
    print("Select movie which you want to delete:")
    show_movie_name()
    movie_to_delete = input(">>>")

    # calling delete functionality
    delete_movie_action(movie_to_delete)


# Movie Delete Logic
def delete_movie_action(movie_to_delete):
    worksheet = data_obj[movie_to_delete]
    data_obj.remove(worksheet)
    print("Movie deleted Successfully")
    data_obj.save('DataSheet.xlsx')
    admin_portal()


# User login portal
def user_login_page(user_login_name):
    print("******Welcome ", user_login_name, " *******")
    # show movie names to user
    # movie_name()
    show_movie_name()
    print("Select movie :")
    user_selected_movie = input(">>>")
    show_movie_details(user_selected_movie)
    print("1. Book Tickets")
    print("2. Cancel Tickets ")
    print("3. Give User Rating  ")
    user_action = int(input("choose your user action: "))
    if (user_action == 1):
        book_movie_ticket_page(user_selected_movie, user_login_name)
    elif (user_action == 2):
        cancel_movie_ticket_page(user_selected_movie, user_login_name)
    elif (user_action == 3):
        user_rating_page(user_selected_movie, user_login_name)
    else:
        print("Choose a valid option")
        home()


# booking page
def book_movie_ticket_page(user_selected_movie, user_login_name):
    # input for booking movie tickets
    print("******Welcome ", user_login_name, " *******")
    show_movie_timings(user_selected_movie)
    print("Select timing")
    user_selected_movie_timing = int(input(">>>"))
    if user_selected_movie_timing == 1:
        user_selected_movie_timing = timings[0]
    elif user_selected_movie_timing == 2:
        user_selected_movie_timing = timings[1]
    elif user_selected_movie_timing == 3:
        user_selected_movie_timing = timings[2]
    else:
        print("Requested timing not available")
        user_login_page(user_login_name)

    print("Select seats")
    user_selected_movie_seats = input(">>>")

    # exception handling for valueError
    try:
        selected_movie_seats = int(user_selected_movie_seats)
        book_movie_ticket(user_selected_movie, user_selected_movie_timing, selected_movie_seats, user_login_name)
        print("Booking successfull")
    except ValueError:
        print("Please enter a numeric value")


# booking tickets login
def book_movie_ticket(user_selected_movie, user_selected_movie_timing, selected_movie_seats, user_login_name):
    worksheet = data_obj[user_selected_movie]
    last_row9 = worksheet.max_row
    for i in range(2, last_row9):
        time_val2 = worksheet.cell(row=i, column=8)
        print(user_selected_movie_timing)
        if time_val2.value == user_selected_movie_timing:
            new_seats = int(worksheet.cell(row=i, column=13).value) - selected_movie_seats
            if (new_seats > 0):
                print("Requested no. of seats are available and being booked")
                print("Seats Remaning", new_seats)
                worksheet.cell(row=i, column=13).value = new_seats
                data_obj.save('DataSheet.xlsx')
                user_login_page(user_login_name)
                break
            else:
                print("Requested no. of seats are more than available seat limit")
                user_login_page(user_login_name)


# cancel tickets page
def cancel_movie_ticket_page(user_selected_movie, user_login_name):
    print("******Welcome ", user_login_name, " *******")
    show_movie_timings(user_selected_movie)
    print("Select timing")
    user_selected_movie_timing = input(">>>")
    print("Select seats")
    user_selected_movie_seats = input(">>>")
    print("Tickets cancelled successfull")
    cancel_movie_ticket(user_selected_movie, user_selected_movie_timing, user_selected_movie_seats)
    user_login_page()


# cancel tickets logic
def cancel_movie_ticket(user_selected_movie, user_selected_movie_timing, user_selected_movie_seats):
    worksheet = data_obj[user_selected_movie]
    last_row9 = worksheet.max_row
    for i in range(2, last_row9):
        time_val2 = worksheet.cell(row=i, column=8)
        if time_val2.value == user_selected_movie_timing:
            new_seats_cancel = int(worksheet.cell(row=i, column=13).value) + int(user_selected_movie_seats)
            worksheet.cell(row=i, column=13).value = new_seats_cancel
            data_obj.save('DataSheet.xlsx')
            break


# showing movie names
def show_movie_name():
    for sheet in data_obj.worksheets:
        print(sheet.title)


# user rating page
def user_rating_page(user_selected_movie, user_login_name):
    print("******Welcome ", user_login_name, " *******")
    print("Please give rating")
    user_selected_movie_rating = input(">>>")
    user_rating(user_selected_movie, user_selected_movie_rating, user_login_name)


# user rating logic
def user_rating(user_selected_movie, user_selected_movie_rating, user_login_name):
    worksheet = data_obj[user_selected_movie]
    current_user_count = int(worksheet.cell(row=2, column=14).value) + 1
    print("user count", current_user_count)
    current_rating = worksheet.cell(row=2, column=6).value
    print("current rating", current_rating)
    new_rating_sum = int(current_rating) + int(user_selected_movie_rating)
    print("new rating sum", new_rating_sum)
    new_rating = (int(new_rating_sum)) / (int(current_user_count))
    print("new rating ", new_rating)
    worksheet.cell(row=2, column=6).value = new_rating
    worksheet.cell(row=2, column=14).value = current_user_count
    data_obj.save('DataSheet.xlsx')
    show_movie_details(user_selected_movie)
    user_login_page(user_login_name)


# return to homepage
home()
